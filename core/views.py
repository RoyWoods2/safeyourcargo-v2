# views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from .models import *
from .forms import *
from django.views.decorators.http import require_POST
from django.contrib.auth.hashers import make_password
from django.db import transaction
from django.http import HttpResponseBadRequest, HttpResponse
from django.core.paginator import Paginator
from weasyprint import HTML
from django.template.loader import render_to_string
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
import requests
from django.utils.formats import date_format
from num2words import num2words
from core.services.facturacion_cl import emitir_factura_exenta_cl_xml, generar_txt_factura_exenta,generar_xml_factura_exenta
from core.services.unlocode_utils import get_ports_by_country,get_airports_by_country,pais_a_codigo # 🔹 IMPORTA AQUÍ
import logging
import tempfile
import io
import os
from django.db import models
from django.core.mail import send_mail
from django.contrib.auth.decorators import user_passes_test, login_required
from django.conf import settings
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.db.models import Count
from django.views.decorators.csrf import csrf_exempt
logger = logging.getLogger(__name__)
from django.contrib.auth import get_user_model
import json
from django.db.models import Q, Count, Sum
from .utils import registrar_actividad, obtener_dolar_observado, descargar_pdf_sii
from xml.etree.ElementTree import Element, SubElement, tostring
from django.views.decorators.http import require_http_methods
from .utils_pdf import generar_pdf_certificado, generar_pdf_factura  
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime,timedelta
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from core.emails import enviar_certificado_y_factura
from django.http import HttpResponseForbidden
from django.utils.html import strip_tags
from .api_client import nsure_api
from django.db import models
from django.views import View
from django.utils import timezone
from django.core.cache import cache
import string
from core.models import UnlocodeEntry
Usuario = get_user_model()

@login_required
def dashboard(request):
    user = request.user

    # ===== Rango de días (default 90) =====
    try:
        days = int(request.GET.get('days', 90))
        days = max(0, min(days, 365))
    except ValueError:
        days = 90
    date_filter = {}
    if days > 0:
        date_filter = {'fecha_creacion__gte': timezone.now() - timedelta(days=days)}

    # ===== Excluir CLIENTES “Prueba*” =====
    TEST_CLIENT_NAMES = ["Prueba", "Prueba 2", "Prueba 3"]
    client_name_filters = Q()
    for n in TEST_CLIENT_NAMES:
        client_name_filters |= Q(nombre__iexact=n)
    test_clients = Cliente.objects.filter(client_name_filters)

    # Tipo de cambio (usa tu función/fallback)
    resultado = obtener_dolar_observado("hans.arancibia@live.com", "Rhad19326366.")
    dolar = Decimal(resultado.get("valor", '950.00'))

    certificados_clientes = []
    certificados_totales = []
    total_certificados_sum = 0
    total_prima_usd = Decimal('0.0')
    total_prima_clp = Decimal('0.0')
    total_clientes = 0
    ultimos_certificados = []
    origen_data = []

    if user.is_superuser or user.rol == "Administrador":
        # Certificados válidos (excluye clientes de prueba + por periodo)
        certificados_ok = (
            CertificadoTransporte.objects
            .exclude(cliente__in=test_clients)
            .filter(**date_filter)
        )

        # Top clientes por cantidad (excluidos “Prueba*”)
        clientes_ok = (
            Cliente.objects
            .exclude(client_name_filters)
            .annotate(total_certificados=Count(
                'certificadotransporte',
                filter=Q(certificadotransporte__in=certificados_ok)
            ))
            .filter(total_certificados__gt=0)
            .order_by('-total_certificados')[:12]
        )
        certificados_clientes = [c.nombre for c in clientes_ok]
        certificados_totales  = [c.total_certificados for c in clientes_ok]
        total_certificados_sum = certificados_ok.count()

        total_prima_usd = certificados_ok.aggregate(
            total=Sum('tipo_mercancia__valor_prima')
        )['total'] or Decimal('0.0')
        total_prima_clp = total_prima_usd * dolar

        total_clientes = (
            Cliente.objects.exclude(client_name_filters)
            .filter(certificadotransporte__in=certificados_ok)
            .distinct().count()
        )

        ultimos_certificados = (
            certificados_ok.select_related('cliente', 'ruta', 'tipo_mercancia')
            .order_by('-id')[:5]
        )

        origen_data = (
            certificados_ok.values('ruta__pais_origen')
            .annotate(cantidad=Count('id'))
            .order_by('-cantidad')
        )

    elif user.rol == "Revendedor":
        propios_certificados = (
            CertificadoTransporte.objects
            .filter(cliente=user.cliente, **date_filter)
            .exclude(cliente__in=test_clients)
        )
        certificados_clientes = [user.cliente.nombre] if user.cliente and not test_clients.filter(pk=user.cliente_id).exists() else []
        certificados_totales  = [propios_certificados.count()] if certificados_clientes else []
        total_certificados_sum = propios_certificados.count()

        total_prima_usd = propios_certificados.aggregate(
            total=Sum('tipo_mercancia__valor_prima')
        )['total'] or Decimal('0.0')
        total_prima_clp = total_prima_usd * dolar

        origen_data = (
            propios_certificados.values('ruta__pais_origen')
            .annotate(cantidad=Count('id')).order_by('-cantidad')
        )

    else:
        propios_certificados = (
            CertificadoTransporte.objects
            .filter(creado_por=user, **date_filter)
            .exclude(cliente__in=test_clients)
        )
        agg = (propios_certificados.values('cliente__nombre')
               .annotate(cantidad=Count('id'))
               .order_by('-cantidad')[:12])
        certificados_clientes = [x['cliente__nombre'] for x in agg]
        certificados_totales  = [x['cantidad'] for x in agg]
        total_certificados_sum = propios_certificados.count()

        total_prima_usd = propios_certificados.aggregate(
            total=Sum('tipo_mercancia__valor_prima')
        )['total'] or Decimal('0.0')
        total_prima_clp = total_prima_usd * dolar

        origen_data = (
            propios_certificados.values('ruta__pais_origen')
            .annotate(cantidad=Count('id')).order_by('-cantidad')
        )

    # Redondeo para UI
    total_prima_usd = total_prima_usd.quantize(Decimal('1.'), rounding=ROUND_HALF_UP)
    total_prima_clp = total_prima_clp.quantize(Decimal('1.'), rounding=ROUND_HALF_UP)

    context = {
        'certificados_clientes': certificados_clientes,
        'certificados_totales': certificados_totales,
        'total_certificados_sum': total_certificados_sum,
        'total_prima_usd': total_prima_usd,
        'total_prima_clp': total_prima_clp,
        'total_clientes': total_clientes,
        'ultimos_certificados': ultimos_certificados,
        'origen_paises': [i['ruta__pais_origen'] for i in origen_data],
        'origen_cantidades': [i['cantidad'] for i in origen_data],
        'days': days,
    }
    return render(request, 'core/dashboard.html', context)



def home_redirect(request):
    return redirect('login')
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('dashboard')  # O la vista que quieras mostrar al entrar
        else:
            messages.error(request, "Usuario o contraseña incorrectos.")

    return render(request, 'core/login.html')
@login_required
def lista_clientes(request):
    if request.user.is_superuser:
        clientes = Cliente.objects.all()
    else:
        if request.user.rol == 'Administrador':
            revendedores = Usuario.objects.filter(creado_por=request.user)
            clientes = Cliente.objects.filter(
                Q(creado_por=request.user) | Q(creado_por__in=revendedores)
            )
        else:
            clientes = Cliente.objects.filter(creado_por=request.user)
    return render(request, 'core/clientes.html', {'clientes': clientes})


@login_required
def form_cliente(request):
    if request.method != 'POST':
        return HttpResponseBadRequest("Solo POST")
    if request.headers.get("x-requested-with", "").lower() != "xmlhttprequest":
        return HttpResponseBadRequest("Solo se acepta AJAX aquí")

    cliente_id = request.POST.get("cliente_id")
    rut = (request.POST.get("rut") or "").strip()
    nombre = (request.POST.get("nombre") or "").strip()

    # Cargar instancia dentro del scope
    instance = None
    if cliente_id:
        if request.user.is_superuser:
            instance = get_object_or_404(Cliente, pk=cliente_id)
        elif request.user.rol == 'Administrador':
            revendedores = Usuario.objects.filter(creado_por=request.user)
            instance = get_object_or_404(
                Cliente,
                Q(creado_por=request.user) | Q(creado_por__in=revendedores),
                pk=cliente_id,
            )
        else:
            instance = get_object_or_404(Cliente, pk=cliente_id, creado_por=request.user)

    # Duplicados por usuario (solo creación, para devolver flags)
    if not cliente_id:
        qs = Cliente.objects.all()
        if not request.user.is_superuser:
            if request.user.rol == 'Administrador':
                revendedores = Usuario.objects.filter(creado_por=request.user)
                qs = qs.filter(Q(creado_por=request.user) | Q(creado_por__in=revendedores))
            else:
                qs = qs.filter(creado_por=request.user)
        if rut and qs.filter(rut__iexact=rut).exists():
            return JsonResponse({'success': False, 'duplicate': True,
                                 'errors': {'rut': ['El RUT ingresado ya existe.']}}, status=400)
        if nombre and qs.filter(nombre__iexact=nombre).exists():
            return JsonResponse({'success': False, 'duplicate_nombre': True,
                                 'errors': {'nombre': ['Ya existe un cliente con este nombre.']}}, status=400)

    form = ClienteForm(request.POST, instance=instance, user=request.user)
    if form.is_valid():
        cliente = form.save(commit=False)
        if not cliente_id:
            cliente.creado_por = request.user
        cliente.save()
        registrar_actividad(request.user, f"{'Editó' if cliente_id else 'Creó'} cliente: {cliente.nombre}")
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'errors': form.errors}, status=400)


@login_required
def editar_cliente(request, pk):
    if request.user.is_superuser:
        cliente = get_object_or_404(Cliente, pk=pk)
    elif request.user.rol == 'Administrador':
        revendedores = Usuario.objects.filter(creado_por=request.user)
        cliente = get_object_or_404(
            Cliente,
            Q(creado_por=request.user) | Q(creado_por__in=revendedores),
            pk=pk,
        )
    else:
        cliente = get_object_or_404(Cliente, pk=pk, creado_por=request.user)

    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente, user=request.user)
        if form.is_valid():
            cliente = form.save()
            registrar_actividad(request.user, f"Editó cliente: {cliente.nombre}")
            return JsonResponse({'success': True})
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    else:
        form = ClienteForm(instance=cliente, user=request.user)
        return render(request, 'core/form_cliente.html', {'form': form})


@require_POST
@login_required
def eliminar_cliente(request, pk):
    if request.user.is_superuser:
        cliente = get_object_or_404(Cliente, pk=pk)
    elif request.user.rol == 'Administrador':
        revendedores = Usuario.objects.filter(creado_por=request.user)
        cliente = get_object_or_404(
            Cliente,
            Q(creado_por=request.user) | Q(creado_por__in=revendedores),
            pk=pk,
        )
    else:
        cliente = get_object_or_404(Cliente, pk=pk, creado_por=request.user)

    nombre_cliente = cliente.nombre
    cliente.delete()
    registrar_actividad(request.user, f"Eliminó cliente: {nombre_cliente}")
    return JsonResponse({'success': True})
# (evita borrar clientes ajenos)  :contentReference[oaicite:10]{index=10}



def obtener_ciudades(request):
    pais_id = request.GET.get('pais_id')
    ciudades = Ciudad.objects.filter(pais_id=pais_id).values('id', 'nombre')
    return JsonResponse(list(ciudades), safe=False)




@login_required
def lista_usuarios(request):
    # ❗️ Esta vista ahora es más simple. Solo necesita pasar los usuarios a la plantilla.
    # La lógica del modal (formularios, clientes, roles) se moverá a get_form_usuario.
    if request.user.is_superuser:
        usuarios = Usuario.objects.filter(is_superuser=False).select_related('cliente')
    elif request.user.rol == "Administrador":
        usuarios_subordinados = Usuario.objects.filter(creado_por=request.user)
        # CORRECCIÓN: Filtrar por `creado_por` es más preciso que por cliente para la jerarquía.
        usuarios = Usuario.objects.filter(
            Q(creado_por=request.user) | Q(creado_por__in=usuarios_subordinados) | Q(pk=request.user.pk)
        ).distinct().select_related('cliente')
    else: # Revendedor y Usuario
        usuarios = Usuario.objects.filter(creado_por=request.user).select_related('cliente')

    return render(request, 'core/usuarios.html', {'usuarios': usuarios})

@login_required
def get_form_usuario(request, user_id=None):
    """
    Devuelve el HTML del formulario para crear o editar un usuario.
    Esta vista es llamada por AJAX desde el modal.
    """
    instance = None
    if user_id:
        # Se asegura de que el usuario que edita tenga permisos sobre el editado
        instance = get_object_or_404(Usuario, pk=user_id)
        # Aquí puedes añadir una capa extra de seguridad para verificar permisos
        # Por ejemplo, un superusuario puede editar a todos, un admin a sus sub-usuarios, etc.

    form = UsuarioForm(instance=instance)
    formset = EmailAdicionalFormSet(instance=instance, prefix='emails')

    # Debes crear esta plantilla parcial
    return render(request, 'core/_form_usuario_content.html', {
        'form': form,
        'email_formset': formset
    })


@require_POST
@login_required
def form_usuario(request):
    if not request.headers.get("x-requested-with") == "XMLHttpRequest":
        return HttpResponseBadRequest("Solo se acepta AJAX aquí")

    usuario_id = request.POST.get("usuario_id")
    instance = None
    if usuario_id:
        instance = get_object_or_404(Usuario, pk=usuario_id)

    form = UsuarioForm(request.POST, instance=instance)
    formset = EmailAdicionalFormSet(request.POST, instance=instance, prefix='emails')

    if form.is_valid() and formset.is_valid():
        try:
            with transaction.atomic():
                # Guarda el formulario principal del usuario
                usuario = form.save(commit=False)
                
                # Asigna la contraseña solo si se proporcionó una nueva
                password = form.cleaned_data.get('password')
                if password:
                    usuario.set_password(password)
                
                # Si es un usuario nuevo, asigna el creador
                if not instance:
                    usuario.creado_por = request.user
                
                usuario.save()

                # Asocia el formset con la instancia del usuario recién guardada y guarda
                formset.instance = usuario
                formset.save()

            registrar_actividad(request.user, f"Guardó datos del usuario: {usuario.username}")
            return JsonResponse({'success': True})
            
        except Exception as e:
            # Captura errores inesperados durante la transacción
            logger.error(f"Error en transacción al guardar usuario: {e}", exc_info=True)
            return JsonResponse({'success': False, 'errors': json.dumps({'__all__': [f"Error inesperado: {str(e)}"]})}, status=400)
    else:
        # Combina errores del formulario principal y del formset
        errors = form.errors.as_data()
        if formset.errors:
            # Añade los errores del formset a la lista para el frontend
            errors['formset'] = formset.errors
        
        # Convertimos a JSON para que el frontend lo pueda procesar
        return JsonResponse({'success': False, 'errors': json.dumps(errors, cls=json.JSONEncoder)}, status=400)


@user_passes_test(lambda u: u.is_superuser)
def aprobar_usuario(request, user_id):
    usuario = get_object_or_404(Usuario, id=user_id, pendiente_aprobacion=True)
    usuario.is_active = True
    usuario.pendiente_aprobacion = False
    usuario.save()
    registrar_actividad(request.user, f"Aprobó al usuario: {usuario.username}")
    send_mail(
        "Tu cuenta fue aprobada",
        f"Hola {usuario.username}, tu cuenta ha sido aprobada. ¡Ya puedes acceder!",
        settings.DEFAULT_FROM_EMAIL,
        [usuario.email]
    )
    
    return JsonResponse({'success': True, 'mensaje': 'Usuario aprobado y notificado.'})


@login_required

@require_POST
def eliminar_usuario(request, pk):
    logger.info(f"Intentando eliminar usuario con ID {pk}")

    try:
        user = Usuario.objects.get(pk=pk)
        user.delete()
        logger.info(f"Usuario con ID {pk} eliminado correctamente.")
        return JsonResponse({'success': True})
    except User.DoesNotExist:
        logger.warning(f"Usuario con ID {pk} no encontrado.")
        return JsonResponse({'success': False, 'error': 'Usuario no encontrado'})
    

@login_required
@require_POST
def toggle_estado_usuario(request, pk):
    try:
        user = Usuario.objects.get(pk=pk)
        user.is_active = not user.is_active
        user.save()
        estado = "activado" if user.is_active else "desactivado"
        registrar_actividad(request.user, f"Cambió estado de usuario: {user.username} a {estado}")

        
        return JsonResponse({
            'success': True,
            'estado': user.is_active,
        })
    except User.DoesNotExist:
        raise Http404("Usuario no encontrado")
    


@login_required
def crear_certificado(request):
    if request.method == 'POST':
        cert_form = CertificadoTransporteForm(request.POST, user=request.user)
        ruta_form = RutaForm(request.POST)
        metodo_form = MetodoEmbarqueForm(request.POST)
        mercancia_form = TipoMercanciaForm(request.POST)
        viaje_form = ViajeForm(request.POST)
        notas_form = NotasNumerosForm(request.POST)

        forms_to_validate = {
            'cert_form': cert_form,
            'ruta_form': ruta_form,
            'metodo_form': metodo_form,
            'mercancia_form': mercancia_form,
            'viaje_form': viaje_form,
            'notas_form': notas_form,
        }

        if all(form.is_valid() for form in forms_to_validate.values()):
            try:
                with transaction.atomic():
                    ruta = ruta_form.save()
                    metodo = metodo_form.save()
                    mercancia = mercancia_form.save()
                    viaje = viaje_form.save(commit=False)

                    # Lógica para asignar FK de País
                    origen_pais = Pais.objects.filter(sigla__iexact=viaje.vuelo_origen_pais).first()
                    destino_pais = Pais.objects.filter(sigla__iexact=viaje.vuelo_destino_pais).first()
                    if origen_pais:
                        viaje.vuelo_origen_pais = origen_pais.nombre
                        viaje.vuelo_origen_pais_fk = origen_pais
                    if destino_pais:
                        viaje.vuelo_destino_pais = destino_pais.nombre
                        viaje.vuelo_destino_pais_fk = destino_pais
                    viaje.save()

                    notas = notas_form.save()

                    certificado = cert_form.save(commit=False)
                    certificado.ruta = ruta
                    certificado.metodo_embarque = metodo
                    certificado.tipo_mercancia = mercancia
                    certificado.viaje = viaje
                    certificado.notas = notas
                    certificado.creado_por = request.user
                    certificado.save()

                    registrar_actividad(request.user, f"Creó certificado: C-{certificado.id}")
                    
                    # --- Generación y emisión de factura ---
                    factura, created = Factura.objects.get_or_create(
                        certificado=certificado,
                        defaults={
                            'numero': Factura.objects.aggregate(max_num=models.Max('numero'))['max_num'] + 1 if Factura.objects.exists() else 1,
                            'razon_social': certificado.cliente.nombre,
                            'rut': certificado.cliente.rut,
                            'direccion': certificado.cliente.direccion,
                            'comuna': certificado.cliente.region or 'Por definir',
                            'ciudad': certificado.cliente.ciudad,
                            'valor_usd': mercancia.valor_prima,
                            'fecha_emision': date.today(),
                            'estado_emision': 'pendiente'
                        }
                    )
                    factura.valor_usd = mercancia.valor_prima
                    
                    resultado_dolar = obtener_dolar_observado(settings.BCCH_USER, settings.BCCH_PASS)
                    dolar = Decimal(resultado_dolar.get("valor", '950.00'))

                    factura.tipo_cambio = dolar
                    factura.valor_clp = (factura.valor_usd or Decimal('0.0')) * dolar
                    
                    # El folio_sii ahora está en None. Se guardará después de obtenerlo de la API.
                    factura.save()

                    # --- ELIMINADAS LAS LÍNEAS QUE ASIGNABAN EL FOLIO MANUALMENTE ---
                    # from .models import obtener_siguiente_folio
                    # factura.folio_sii = obtener_siguiente_folio()
                    # factura.save()
                    # -----------------------------------------------------------------

                    response_facturacion = emitir_factura_exenta_cl_xml(factura)
                    
                    if response_facturacion and response_facturacion.get('success'):
                        # El folio y la URL se obtienen de la respuesta de la API y se guardan aquí.
                        factura.folio_sii = response_facturacion.get('folio_sii')
                        factura.url_pdf_sii = response_facturacion.get('url_pdf_sii')
                        factura.estado_emision = 'exito'
                        messages.success(request, "Certificado y factura emitida correctamente con timbre SII.")
                        logger.info(f"Factura {factura.id} emitida con Folio SII: {factura.folio_sii}, URL: {factura.url_pdf_sii}")
                    else:
                        error_detalle = response_facturacion.get('error', 'Error desconocido al emitir a facturacion.cl')
                        factura.estado_emision = 'fallida'
                        factura.observaciones = f"Error al emitir DTE: {error_detalle}"
                        messages.error(request, f"Certificado creado, pero error al emitir factura electrónica: {error_detalle}")
                        logger.error(f"Error al emitir factura {factura.id} a facturacion.cl: {error_detalle}")
                    
                    # Guardamos la factura por segunda vez para almacenar el folio y estado final.
                    factura.save()

                    # --- Lógica de Envío de Correos (sin cambios) ---
                    try:
                        certificado_pdf_buffer = generar_pdf_certificado(certificado, request)
                        factura_pdf_buffer = generar_pdf_factura(certificado, request) 

                        otros_emails_manuales_str = cert_form.cleaned_data.get('otros_emails_copia', '')
                        emails_manuales = [e.strip() for e in (otros_emails_manuales_str or '').split(',') if e.strip()]
                        emails_guardados_usuario = request.user.get_lista_emails_adicionales()
                        todos_los_destinatarios_extra = list(set(emails_manuales + emails_guardados_usuario))
                        
                        enviar_certificado_y_factura(
                            certificado=certificado,
                            pdf_cert=certificado_pdf_buffer,
                            factura_obj=factura,
                            pdf_fact=factura_pdf_buffer,
                            destinatarios_extra=todos_los_destinatarios_extra
                        )
                        logger.info(f"Proceso de envío de correo para certificado C-{certificado.id} iniciado.")

                    except Exception as e:
                        logger.error(f"Error general en el proceso de envío de correo para el certificado C-{certificado.id}: {e}", exc_info=True)
                        messages.warning(request, "El certificado se creó y la factura se procesó, pero hubo un problema al enviar la notificación por correo.")
                    
                    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': True,
                            'factura_emitida': response_facturacion.get("success", False),
                            'resultado': response_facturacion
                        })

                    return redirect('crear_certificado')

            except Exception as e:
                logger.error(f"Error general en la transacción de creación de certificado/factura: {e}", exc_info=True)
                messages.error(request, f"Error al generar el certificado o la factura: {e}")
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'errors': {'__all__': [str(e)]}}, status=400)
                return redirect('crear_certificado')

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            errors = {form_name: form_instance.errors for form_name, form_instance in forms_to_validate.items() if not form_instance.is_valid()}
            logger.error(f"Errores de validación de formulario: {errors}")
            return JsonResponse({
                'success': False,
                'errors': errors
            }, status=400)

    # --- El resto de la vista (lógica GET) se mantiene igual ---
    fecha_inicio = request.GET.get('inicio')
    fecha_fin = request.GET.get('fin')
    busqueda = request.GET.get('q', '').strip()

    certificados_list = CertificadoTransporte.objects.select_related('cliente', 'notas', 'creado_por').order_by('-id')

    if not request.user.is_superuser:
        certificados_list = certificados_list.filter(creado_por=request.user)

    if fecha_inicio:
        try:
            certificados_list = certificados_list.filter(fecha_creacion__gte=datetime.strptime(fecha_inicio, '%Y-%m-%d').date())
        except ValueError:
            pass

    if fecha_fin:
        try:
            certificados_list = certificados_list.filter(fecha_creacion__lte=datetime.strptime(fecha_fin, '%Y-%m-%d').date())
        except ValueError:
            pass

    if busqueda:
        certificados_list = certificados_list.filter(
            Q(cliente__nombre__icontains=busqueda) |
            Q(cliente__rut__icontains=busqueda) |
            Q(id__icontains=busqueda) |
            Q(notas__numero_factura__icontains=busqueda)
        )

    paginator = Paginator(certificados_list, 10)
    certificados = paginator.get_page(request.GET.get('page'))

    context = {
        'cert_form': CertificadoTransporteForm(user=request.user),
        'ruta_form': RutaForm(),
        'metodo_form': MetodoEmbarqueForm(),
        'mercancia_form': TipoMercanciaForm(),
        'viaje_form': ViajeForm(),
        'notas_form': NotasNumerosForm(),
        'certificados': certificados,
        'filtros': {
            'inicio': fecha_inicio or '',
            'fin': fecha_fin or '',
            'q': busqueda,
        }
    }

    return render(request, 'certificados/crear_certificado.html', context)


@login_required
def clean_valor_prima(self):
    prima = self.cleaned_data.get('valor_prima')
    asegurado = (self.cleaned_data.get('valor_fca', 0) + self.cleaned_data.get('valor_flete', 0)) * 1.10
    minimo = max(asegurado * 0.0015, 20.0)

    if prima < minimo:
        raise forms.ValidationError(f"La prima debe ser al menos ${minimo:.2f}")
    return prima

@login_required
def certificado_pdf(request, pk):
    certificado = get_object_or_404(CertificadoTransporte, pk=pk)

    html_string = render_to_string('certificados/certificado_pdf.html', {
        'certificado': certificado,
    })

    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf = html.write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'filename="certificado-C{certificado.id}.pdf"'
    return response

@login_required
def factura_pdf(request, pk):
    from decimal import Decimal
    from datetime import date
    import requests
    from num2words import num2words
    from django.utils.formats import date_format

    certificado = get_object_or_404(CertificadoTransporte, pk=pk)

    # Obtener o crear la factura usando el valor prima real de la mercancía (no recalculado)
    factura, created = Factura.objects.get_or_create(
        certificado=certificado,
        defaults={
            'numero': Factura.objects.count() + 1,
            'razon_social': certificado.cliente.nombre,
            'rut': certificado.cliente.rut,
            'direccion': certificado.cliente.direccion,
            'comuna': certificado.cliente.region or 'Por definir',
            'ciudad': certificado.cliente.ciudad,
            'valor_usd': certificado.tipo_mercancia.valor_prima,  # ✅ valor prima real
            'fecha_emision': date.today()
        }
    )
    factura.valor_usd = certificado.tipo_mercancia.valor_prima  # ✅ actualiza el valor prima real

    # Obtener tipo de cambio dinámico
    resultado = obtener_dolar_observado("hans.arancibia@live.com", "Rhad19326366.")
    if "valor" in resultado:
        dolar = Decimal(resultado["valor"])
    else:
        dolar = Decimal('950.00')


    # Calcular valor CLP
    factura.tipo_cambio = dolar
    factura.valor_clp = (factura.valor_usd or Decimal('0.0')) * dolar
    factura.save()

    # Calcular total en palabras
    total_palabras = num2words(int(factura.valor_clp), lang='es').replace("coma cero cero", "")

    # Formatear fecha
    fecha_formateada = date_format(factura.fecha_emision, "d \d\e F \d\e Y")

    # Renderizar el template a HTML
    html_string = render_to_string('certificados/factura_pdf.html', {
        'factura': factura,
        'total_palabras': total_palabras,
        'fecha_formateada': fecha_formateada,
    })

    # Convertir HTML a PDF
    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf_file = html.write_pdf()

    # Devolver respuesta PDF
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="factura-C{certificado.id}.pdf"'
    return response



@login_required
def factura_confirmacion(request, pk):
    factura = get_object_or_404(Factura, pk=pk)
    resultado = request.session.pop('factura_emitida_resultado', None)

    return render(request, 'certificados/factura_confirmacion.html', {
        'factura': factura,
        'resultado': resultado
    })


from django.http import HttpResponseForbidden

@login_required
def vista_cobranzas(request):
    current_user = request.user

    # Inicia con todas las cobranzas
    cobranzas = Cobranza.objects.select_related(
        'certificado', 'certificado__cliente', 'certificado__metodo_embarque', 'certificado__factura', 'certificado__notas'
    )

    # Excepción para superadministradores: Si el usuario es superadmin, ve todos los certificados.
    if current_user.is_superuser:
        # No se aplica ningún filtro de usuario; el queryset ya contiene todos los objetos.
        pass
    else:
        # Filtrado basado en la jerarquía de usuario para usuarios no superadmin
        if current_user.rol == 'Administrador':
            # Un administrador regular (no superadmin) puede ver todos los certificados.
            # Esto asume que 'Administrador' no es superadmin, o si es superadmin, ya lo manejamos arriba.
            # Si un 'Administrador' regular solo debe ver sus propios y sub-usuarios, ajustar aquí.
            # Por ahora, se mantiene como "ver todo" para el rol 'Administrador' si no es superuser.
            pass
        elif current_user.rol == 'Revendedor':
            # Un revendedor puede ver certificados creados por ellos y por los usuarios que crearon.
            sub_user_ids = Usuario.objects.filter(creado_por=current_user).values_list('id', flat=True)
            allowed_user_ids = list(sub_user_ids) + [current_user.id]
            cobranzas = cobranzas.filter(certificado__creado_por__id__in=allowed_user_ids)
        else: # Rol 'Usuario' por defecto o cualquier otro
            # Un usuario regular solo puede ver los certificados que él mismo creó.
            cobranzas = cobranzas.filter(certificado__creado_por=current_user)

    # Aplica los filtros existentes de la solicitud GET
    cliente = request.GET.get("cliente")
    rut = request.GET.get("rut")
    certificado_id = request.GET.get("certificado")
    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")

    if cliente:
        cobranzas = cobranzas.filter(certificado__cliente__nombre__icontains=cliente)
    if rut:
        cobranzas = cobranzas.filter(certificado__cliente__rut__icontains=rut)
    if certificado_id:
        try:
            certificado_id = int(certificado_id)
            cobranzas = cobranzas.filter(certificado__id=certificado_id)
        except ValueError:
            cobranzas = cobranzas.none()
    if inicio:
        try:
            fecha_inicio = datetime.strptime(inicio, "%Y-%m-%d").date()
            cobranzas = cobranzas.filter(certificado__fecha_creacion__gte=fecha_inicio)
        except ValueError:
            pass
    if fin:
        try:
            fecha_fin = datetime.strptime(fin, "%Y-%m-%d").date()
            cobranzas = cobranzas.filter(certificado__fecha_creacion__lt=fecha_fin + timedelta(days=1))
        except ValueError:
            pass
            
    # Ordena los resultados
    cobranzas = cobranzas.order_by('-certificado__fecha_creacion') # Cambiado a fecha_creacion para consistencia

    filtros = {
        "cliente": cliente or "",
        "rut": rut or "",
        "certificado": certificado_id if certificado_id is not None else "",
        "inicio": inicio or "",
        "fin": fin or "",
    }

    return render(request, 'core/cobranzas.html', {
        'cobranzas': cobranzas,
        'filtros': filtros,
    })


@login_required
def generar_pdf_cobranza(request, certificado_id):
    cobro = Cobranza.objects.select_related('certificado', 'certificado__cliente', 'certificado__metodo_embarque').get(certificado_id=certificado_id)

    # Ruta absoluta del logo para WeasyPrint
    logo_path = os.path.join(settings.BASE_DIR, 'core', 'static', 'img', 'safe_logo.png')
    logo_fixed = logo_path.replace("\\", "/")
    html_string = render_to_string('core/pdf_cobranza.html', {
        'cobro': cobro,
        
        'logo_path': f'file:///{logo_fixed}',

    })

    pdf_file = io.BytesIO()
    HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf(pdf_file)

    response = HttpResponse(pdf_file.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename=cobranza_C{cobro.certificado.id}.pdf'
    return response


# ---------- Helper de normalización ----------
ISO_ALIASES = {
    # agrega más alias si lo necesitas
    "united states":        {"iso2": "US", "iso3": "USA", "names": ["United States", "United States of America", "USA", "U.S.", "U.S.A."]},
    "russia":               {"iso2": "RU", "iso3": "RUS", "names": ["Russia", "Russian Federation"]},
    "south korea":          {"iso2": "KR", "iso3": "KOR", "names": ["South Korea", "Republic of Korea", "Korea, Republic of"]},
    "czech republic":       {"iso2": "CZ", "iso3": "CZE", "names": ["Czech Republic", "Czechia"]},
    "uk":                   {"iso2": "GB", "iso3": "GBR", "names": ["United Kingdom", "Great Britain", "UK"]},
    "uae":                  {"iso2": "AE", "iso3": "ARE", "names": ["United Arab Emirates", "UAE"]},
}

def normalize_country(pais: str, pais_code: str):
    p = (pais or "").strip()
    c = (pais_code or "").strip()
    iso2 = ""
    iso3 = ""
    names = []

    # si viene código
    if len(c) == 2:
        iso2 = c.upper()
    elif len(c) == 3:
        iso3 = c.upper()

    # si viene nombre, busca alias
    key = p.lower()
    alias = ISO_ALIASES.get(key)
    if alias:
        iso2 = iso2 or alias["iso2"]
        iso3 = iso3 or alias["iso3"]
        names = alias["names"][:]
    elif p:
        names = [p]  # intenta con el nombre tal cual

    # por si mandaron el nombre en código
    if not iso2 and len(p) == 2:
        iso2 = p.upper()
    if not iso3 and len(p) == 3:
        iso3 = p.upper()

    return iso2, iso3, names
NSURE_API = "https://igi.nsure.net/api/seaports/search/term/{term}"
NSURE_HEADERS = {
    "Content-Type": "application/json",
    "Accept": "application/json",
    "X-NS-API-Key": "PqrED4vo2UI8I8TlPTgOKmCo0C1OuSUzbgbpIBHQwgnA2343xbtwqjhmGZ2bckvp",
}

def _nsure_ports_fallback_by_country(iso2: str, names=None, timeout=10, per_term_limit=200):
    """
    Busca puertos por letras/dígitos y filtra por país usando:
    - countryCode == iso2  OR
    - countryName ∈ names  OR
    - locode empieza con iso2
    Deduplica por locode. Cachea 24h.
    """
    iso2 = (iso2 or "").upper().strip()
    accepted_names = {n.lower() for n in (names or []) if n}

    if not iso2 or len(iso2) != 2:
        return []

    cache_key = f"ports_nsure:{iso2}"
    cached = cache.get(cache_key)
    if cached:
        return cached

    terms = list(string.ascii_lowercase) + list("0123456789")
    seen = set()
    out = []

    for t in terms:
        try:
            r = requests.get(NSURE_API.format(term=t), headers=NSURE_HEADERS, timeout=timeout)
            r.raise_for_status()
            data = r.json()
            items = data if isinstance(data, list) else data.get("results", []) or []

            for it in items[:per_term_limit]:
                name_field = (it.get("countryName") or it.get("country_name") or it.get("country") or "").strip()
                code_field = (it.get("countryCode") or it.get("country_code") or "").upper().strip()
                locode     = (it.get("locode") or it.get("code") or "").upper().strip()
                name_port  = (it.get("name") or it.get("portName") or it.get("port") or "").strip()

                # --- 3 caminos de aceptación ---
                ok = False
                if code_field == iso2:
                    ok = True
                elif name_field and name_field.lower() in accepted_names:
                    ok = True
                elif locode.startswith(iso2):
                    ok = True
                if not ok:
                    continue

                if not locode or not name_port:
                    continue
                if locode in seen:
                    continue
                seen.add(locode)
                out.append({"name": name_port, "locode": locode, "countryCode": code_field or iso2})
        except requests.RequestException:
            continue

    cache.set(cache_key, out, 60 * 60 * 24)
    return out

@csrf_exempt
def obtener_ciudades(request):
    if request.method == "POST":
        import json
        datos = json.loads(request.body)
        pais = datos.get("pais")

        # 👉 La API externa espera GET, no POST
        response = requests.get(
            f"https://countriesnow.space/api/v0.1/countries/cities/q?country={pais}"
        )

        if response.ok:
            data = response.json()
            return JsonResponse({"ciudades": data.get("data", [])})
        return JsonResponse({"error": "No se pudieron obtener las ciudades"}, status=500)
    
    return JsonResponse({"error": "Método no permitido"}, status=405)
# ---------- Aeropuertos ----------
@csrf_exempt
def obtener_aeropuertos(request):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)
    try:
        data = json.loads(request.body)
        iso2, iso3, names = normalize_country(data.get("pais"), data.get("pais_code"))

        aeropuertos = []

        # 1) ISO-2
        if iso2:
            aeropuertos = get_airports_by_country(iso2)
        # 2) ISO-3
        if not aeropuertos and iso3:
            aeropuertos = get_airports_by_country(iso3)
        # 3) Aliases de nombre
        if not aeropuertos:
            for nm in names:
                if not nm:
                    continue
                aeropuertos = get_airports_by_country(nm)
                if aeropuertos:
                    break

        return JsonResponse({"aeropuertos": aeropuertos})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

def obtener_paises(request):
    url = "https://restcountries.com/v3.1/all?fields=name,cca2,flags"
    response = requests.get(url)
    if response.ok:
        data = response.json()
        paises = sorted([{
            "nombre": c["name"]["common"],
            "codigo": c["cca2"],
            "bandera": c["flags"]["svg"]
        } for c in data], key=lambda x: x["nombre"])
        return JsonResponse({"paises": paises})
    return JsonResponse({"error": "No se pudieron obtener los países"}, status=500)



@csrf_exempt
def obtener_ciudades(request):
    if request.method == "POST":
        datos = json.loads(request.body)
        pais = datos.get("pais")

        # API correcta: requiere POST con JSON {"country": "Chile"}
        url = "https://countriesnow.space/api/v0.1/countries/cities"
        response = requests.post(url, json={"country": pais})

        if response.ok:
            data = response.json()
            return JsonResponse({"ciudades": data.get("data", [])})
        return JsonResponse({"error": "No se pudieron obtener las ciudades"}, status=500)
    
    return JsonResponse({"error": "Método no permitido"}, status=405)

# ---------- Puertos (UN/LOCODE) ----------
@csrf_exempt
def obtener_unlocode(request):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)
    try:
        data = json.loads(request.body or "{}")
        # acepta ambos, pero prioriza pais_code
        pais_code = (data.get("pais_code") or data.get("pais") or "").strip().upper()
        funcion   = str(data.get("function", "1")).strip()
        q         = (data.get("q") or "").strip()

        if len(pais_code) != 2:
            return JsonResponse({"ubicaciones": []})

        cache_key = f"unlocode:{pais_code}:{'1' in funcion}:{q.lower()}"
        cached = cache.get(cache_key)
        if cached is not None:
            return JsonResponse({"ubicaciones": cached})

        qs = UnlocodeEntry.objects.filter(country_code=pais_code)
        if '1' in funcion:  # puertos marítimos
            qs = qs.filter(function__icontains='1')
        if q:
            qs = qs.filter(
                Q(name__icontains=q) |
                Q(locode__icontains=q) |
                Q(subdiv__icontains=q)
            )

        qs = qs.order_by("name")[:2000]  # seguridad para no traer todo
        ubicaciones = [
            {
                "name": u.name,
                "locode": u.locode,
                "countryCode": u.country_code,
                "subdiv": u.subdiv or ""
            }
            for u in qs
        ]
        cache.set(cache_key, ubicaciones, 60 * 60 * 12)  # 12 horas
        return JsonResponse({"ubicaciones": ubicaciones})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)





@login_required
def obtener_clientes_disponibles(request):
    if request.user.is_superuser:
        clientes = Cliente.objects.all()
    elif request.user.rol == "Administrador":
        usuarios_subordinados = Usuario.objects.filter(cliente=request.user.cliente, rol="Revendedor")
        clientes_ids = [request.user.cliente.id] + list(usuarios_subordinados.values_list('cliente', flat=True))
        clientes = Cliente.objects.filter(id__in=clientes_ids)
    else:
        clientes = Cliente.objects.filter(id=request.user.cliente_id)

    data = [{'id': c.id, 'nombre': c.nombre} for c in clientes]
    return JsonResponse({'clientes': data})
@login_required
def obtener_datos_modal_usuario(request):
    if not request.headers.get("x-requested-with") == "XMLHttpRequest":
        return HttpResponseBadRequest("Solo se acepta AJAX aquí")

    # CLIENTES según el rol actual
    if request.user.is_superuser:
        clientes = Cliente.objects.all()
    elif request.user.rol == "Administrador":
        usuarios_sub = Usuario.objects.filter(cliente=request.user.cliente, rol="Revendedor")
        clientes_ids = [request.user.cliente.id] + list(usuarios_sub.values_list('cliente', flat=True))
        clientes = Cliente.objects.filter(id__in=clientes_ids)
    else:
        clientes = Cliente.objects.filter(id=request.user.cliente_id)

    clientes_data = [{'id': c.id, 'nombre': c.nombre} for c in clientes]

    # ROLES permitidos
    if request.user.is_superuser:
        roles = ['Usuario', 'Administrador', 'Revendedor']
    else:
        roles = ['Revendedor']

    return JsonResponse({'clientes': clientes_data, 'roles': roles})


@login_required
def obtener_logs_actividad(request):
    # Filtra los logs según la jerarquía
    if request.user.is_superuser:
        logs = LogActividad.objects.all()
    elif request.user.rol == "Administrador":
        usuarios_sub = Usuario.objects.filter(creado_por=request.user)
        logs = LogActividad.objects.filter(
            usuario__in=[request.user] + list(usuarios_sub)
        )
    elif request.user.rol == "Revendedor":
        logs = LogActividad.objects.filter(usuario=request.user)
    else:
        logs = LogActividad.objects.none()

    logs_data = [{
        'usuario': log.usuario.username,
        'mensaje': log.mensaje,
        'fecha': log.fecha.strftime("%d-%m-%Y %H:%M")
    } for log in logs.order_by('-fecha')[:15]]

    return JsonResponse({'logs': logs_data})


@login_required
def probar_envio_factura(request):
    logs = []
    try:
        logs.append("🔍 Iniciando prueba de envío de factura...")

        # ✅ Usar el último certificado correctamente ordenado por ID
        certificado = CertificadoTransporte.objects.select_related('cliente', 'ruta', 'tipo_mercancia').order_by('-id').first()
        if not certificado:
            messages.error(request, "❌ No hay certificados disponibles para simular el envío.")
            return redirect('crear_certificado')

        logs.append(f"✅ Certificado seleccionado: C-{certificado.id}")

        # Crear o recuperar factura asociada
        factura, created = Factura.objects.get_or_create(
            certificado=certificado,
            defaults={
                'numero': Factura.objects.count() + 1,
                'razon_social': certificado.cliente.nombre,
                'rut': certificado.cliente.rut,
                'direccion': certificado.cliente.direccion,
                'comuna': certificado.cliente.region or 'Por definir',
                'ciudad': certificado.cliente.ciudad,
                'valor_usd': certificado.tipo_mercancia.valor_prima,
                'fecha_emision': date.today()
            }
        )

        # Reasignar valor prima por seguridad
        factura.valor_usd = certificado.tipo_mercancia.valor_prima

        # Obtener tipo de cambio actual
        resultado = obtener_dolar_observado("hans.arancibia@live.com", "Rhad19326366.")
        dolar = Decimal(resultado.get("valor", '950.00'))

        factura.tipo_cambio = dolar
        factura.valor_clp = (factura.valor_usd or Decimal('0.0')) * dolar
        factura.save()

        logs.append(f"📦 Factura generada (USD {factura.valor_usd} → CLP {factura.valor_clp})")

        # Enviar a facturacion.cl
        resultado = emitir_factura_exenta_cl_xml(factura)
        if resultado.get('success'):
            logs.append("✅ ENVÍO EXITOSO a Facturacion.cl")
            logs.append(f"📨 Respuesta: {resultado['respuesta']}")
        else:
            logs.append("❌ ERROR en el envío a Facturacion.cl")
            logs.append(f"🧨 Detalle: {resultado.get('error')}")

    except Exception as ex:
        logs.append("❌ Excepción no controlada:")
        logs.append(str(ex))

    return render(request, 'certificados/prueba_envio_resultado.html', {'logs': logs})



@login_required
def probar_envio(request):
    try:
        factura = Factura.objects.select_related('certificado').order_by('-id').first()

        if not factura:
            return HttpResponse("No hay facturas disponibles para prueba.", status=404)

        contenido_xml = generar_xml_factura_exenta(factura)
        filename = f"factura_exenta_C{factura.certificado.id}.xml"

        response = HttpResponse(contenido_xml, content_type='application/xml')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        return HttpResponse(f"❌ Error al generar archivo XML de prueba: {str(e)}", status=500)

    
    
def descargar_xml_dte(request):
    factura = Factura.objects.select_related('certificado', 'certificado__cliente', 'certificado__ruta').last()
    if not factura:
        return HttpResponse("No hay factura para generar XML", status=400)

    certificado = factura.certificado
    cliente = certificado.cliente

    # XML base
    root = Element("DTE", version="1.0")
    doc = SubElement(root, "Documento", ID="F1T34")

    encabezado = SubElement(doc, "Encabezado")
    id_doc = SubElement(encabezado, "IdDoc")
    SubElement(id_doc, "TipoDTE").text = "34"
    SubElement(id_doc, "Folio").text = "1"
    SubElement(id_doc, "FchEmis").text = factura.fecha_emision.strftime("%Y-%m-%d")

    emisor = SubElement(encabezado, "Emisor")
    SubElement(emisor, "RUTEmisor").text = "76000555-0"
    SubElement(emisor, "RznSoc").text = "Roberto Gomez"
    SubElement(emisor, "GiroEmis").text = "Importación"
    SubElement(emisor, "Acteco").text = "515009"
    SubElement(emisor, "DirOrigen").text = "Pedro de Valdivia 25"
    SubElement(emisor, "CmnaOrigen").text = "Providencia"
    SubElement(emisor, "CiudadOrigen").text = "SANTIAGO"

    receptor = SubElement(encabezado, "Receptor")
    SubElement(receptor, "RUTRecep").text = cliente.rut or "11111111-1"
    SubElement(receptor, "CdgIntRecep").text = "123123123"
    SubElement(receptor, "RznSocRecep").text = cliente.nombre or "CLIENTE"
    SubElement(receptor, "GiroRecep").text = "SERVICIO"
    SubElement(receptor, "DirRecep").text = cliente.direccion or "SIN DIRECCIÓN"
    SubElement(receptor, "CmnaRecep").text = cliente.region or "SANTIAGO"
    SubElement(receptor, "CiudadRecep").text = cliente.ciudad or "SANTIAGO"

    totales = SubElement(encabezado, "Totales")
    total_valor = int(factura.valor_clp or 0)
    SubElement(totales, "MntExe").text = str(total_valor)
    SubElement(totales, "MntTotal").text = str(total_valor)

    detalle = SubElement(doc, "Detalle")
    SubElement(detalle, "NroLinDet").text = "1"
    cdg_item = SubElement(detalle, "CdgItem")
    SubElement(cdg_item, "TpoCodigo").text = "INT1"
    SubElement(cdg_item, "VlrCodigo").text = "PE"
    SubElement(detalle, "IndExe").text = "1"
    SubElement(detalle, "NmbItem").text = "PRODUCTO EXENTO"
    SubElement(detalle, "QtyItem").text = "1"
    SubElement(detalle, "UnmdItem").text = "UN"
    SubElement(detalle, "PrcItem").text = str(total_valor)
    SubElement(detalle, "MontoItem").text = str(total_valor)

    xml_string = tostring(root, encoding="utf-8", method="xml")

    # Devolver XML como descarga
    response = HttpResponse(xml_string, content_type='application/xml')
    response['Content-Disposition'] = 'attachment; filename=DTE_factura_exenta.xml'
    return response



@require_http_methods(["GET"])
def buscar_aeronaves(request):
    """
    Buscar aerolíneas con múltiples APIs y fallback robusto
    """
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    # Intentar con diferentes APIs
    results = []
    
    # 1. Intentar con AviationStack (si funciona)
    try:
        results = buscar_aeronaves_aviationstack(query)
        if results:
            return JsonResponse({
                'results': results[:10],
                'source': 'AviationStack API',
                'total': len(results)
            })
    except Exception as e:
        print(f"AviationStack falló: {e}")
    
    # 2. Intentar con API alternativa (OpenSky Network - gratuita)
    try:
        results = buscar_aeronaves_opensky(query)
        if results:
            return JsonResponse({
                'results': results[:10],
                'source': 'OpenSky Network API',
                'total': len(results)
            })
    except Exception as e:
        print(f"OpenSky falló: {e}")
    
    # 3. Usar fallback con base de datos local
    results = buscar_aeronaves_fallback(query)
    
    return JsonResponse({
        'results': results[:10],
        'source': 'Fallback Database',
        'total': len(results)
    })


def buscar_aeronaves_aviationstack(query):
    """
    AviationStack API - Solo si funciona
    """
    try:
        api_key = "9b58152127ed762a0bb0f7165d17ce20"
        
        # Intentar endpoint de airlines
        url = f"https://api.aviationstack.com/v1/airlines"
        params = {
            'access_key': api_key,
            'search': query,
            'limit': 10
        }
        
        response = requests.get(url, params=params, timeout=8)
        
        if response.status_code == 200:
            data = response.json()
            results = []
            
            if 'data' in data and data['data']:
                for airline in data.get('data', []):
                    airline_name = airline.get('airline_name', '')
                    iata_code = airline.get('iata_code', '')
                    icao_code = airline.get('icao_code', '')
                    
                    if airline_name:
                        display_name = airline_name
                        if iata_code:
                            display_name += f" ({iata_code})"
                        elif icao_code:
                            display_name += f" ({icao_code})"
                        
                        results.append({
                            'id': iata_code or icao_code or airline_name.replace(' ', '_').lower(),
                            'name': display_name,
                            'type': 'airline',
                            'iata_code': iata_code,
                            'icao_code': icao_code
                        })
            
            return results
        else:
            print(f"AviationStack error: {response.status_code} - {response.text}")
            return []
            
    except Exception as e:
        print(f"Error AviationStack: {e}")
        return []


def buscar_aeronaves_opensky(query):
    """
    OpenSky Network API - COMPLETAMENTE GRATUITA
    """
    try:
        # OpenSky Network tiene una API pública sin necesidad de API key
        # Endpoint para obtener información de vuelos y aerolíneas
        url = "https://opensky-network.org/api/flights/all"
        
        # Parámetros para buscar vuelos recientes (últimas 2 horas)
        import time
        current_time = int(time.time())
        two_hours_ago = current_time - (2 * 3600)
        
        params = {
            'begin': two_hours_ago,
            'end': current_time
        }
        
        response = requests.get(url, params=params, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            results = []
            query_upper = query.upper()
            seen_callsigns = set()
            
            # Procesar vuelos para extraer callsigns de aerolíneas
            for flight in data:
                if len(flight) > 1 and flight[1]:  # callsign está en índice 1
                    callsign = flight[1].strip()
                    if callsign and query_upper in callsign.upper():
                        if callsign not in seen_callsigns:
                            seen_callsigns.add(callsign)
                            
                            # Intentar extraer código de aerolínea del callsign
                            airline_code = callsign[:3] if len(callsign) >= 3 else callsign
                            
                            results.append({
                                'id': callsign.replace(' ', '_').lower(),
                                'name': f"{callsign} (Vuelo Activo)",
                                'type': 'airline',
                                'callsign': callsign,
                                'airline_code': airline_code
                            })
                            
                            if len(results) >= 15:  # Limitar resultados
                                break
            
            return results
        else:
            print(f"OpenSky error: {response.status_code}")
            return []
            
    except Exception as e:
        print(f"Error OpenSky: {e}")
        return []


def buscar_aeronaves_fallback(query):
    """
    Fallback con base de datos extensa de aerolíneas
    """
    # Base de datos completa de aerolíneas principales
    aerolineas_db = [
        # Aerolíneas principales internacionales
        {"name": "American Airlines", "iata": "AA", "icao": "AAL", "country": "USA"},
        {"name": "Delta Air Lines", "iata": "DL", "icao": "DAL", "country": "USA"},
        {"name": "United Airlines", "iata": "UA", "icao": "UAL", "country": "USA"},
        {"name": "Southwest Airlines", "iata": "WN", "icao": "SWA", "country": "USA"},
        {"name": "JetBlue Airways", "iata": "B6", "icao": "JBU", "country": "USA"},
        {"name": "Alaska Airlines", "iata": "AS", "icao": "ASA", "country": "USA"},
        
        # Aerolíneas europeas
        {"name": "Lufthansa", "iata": "LH", "icao": "DLH", "country": "Germany"},
        {"name": "British Airways", "iata": "BA", "icao": "BAW", "country": "UK"},
        {"name": "Air France", "iata": "AF", "icao": "AFR", "country": "France"},
        {"name": "KLM Royal Dutch Airlines", "iata": "KL", "icao": "KLM", "country": "Netherlands"},
        {"name": "Iberia", "iata": "IB", "icao": "IBE", "country": "Spain"},
        {"name": "Alitalia", "iata": "AZ", "icao": "AZA", "country": "Italy"},
        {"name": "Swiss International Air Lines", "iata": "LX", "icao": "SWR", "country": "Switzerland"},
        {"name": "Austrian Airlines", "iata": "OS", "icao": "AUA", "country": "Austria"},
        {"name": "SAS Scandinavian Airlines", "iata": "SK", "icao": "SAS", "country": "Sweden"},
        {"name": "Finnair", "iata": "AY", "icao": "FIN", "country": "Finland"},
        {"name": "TAP Air Portugal", "iata": "TP", "icao": "TAP", "country": "Portugal"},
        {"name": "Aer Lingus", "iata": "EI", "icao": "EIN", "country": "Ireland"},
        {"name": "Ryanair", "iata": "FR", "icao": "RYR", "country": "Ireland"},
        {"name": "EasyJet", "iata": "U2", "icao": "EZY", "country": "UK"},
        
        # Aerolíneas asiáticas
        {"name": "Singapore Airlines", "iata": "SQ", "icao": "SIA", "country": "Singapore"},
        {"name": "Cathay Pacific", "iata": "CX", "icao": "CPA", "country": "Hong Kong"},
        {"name": "Japan Airlines", "iata": "JL", "icao": "JAL", "country": "Japan"},
        {"name": "All Nippon Airways", "iata": "NH", "icao": "ANA", "country": "Japan"},
        {"name": "Korean Air", "iata": "KE", "icao": "KAL", "country": "South Korea"},
        {"name": "Asiana Airlines", "iata": "OZ", "icao": "AAR", "country": "South Korea"},
        {"name": "China Eastern Airlines", "iata": "MU", "icao": "CES", "country": "China"},
        {"name": "China Southern Airlines", "iata": "CZ", "icao": "CSN", "country": "China"},
        {"name": "Air China", "iata": "CA", "icao": "CCA", "country": "China"},
        {"name": "Thai Airways", "iata": "TG", "icao": "THA", "country": "Thailand"},
        {"name": "Malaysia Airlines", "iata": "MH", "icao": "MAS", "country": "Malaysia"},
        {"name": "Philippine Airlines", "iata": "PR", "icao": "PAL", "country": "Philippines"},
        {"name": "Cebu Pacific", "iata": "5J", "icao": "CEB", "country": "Philippines"},
        
        # Aerolíneas de Medio Oriente
        {"name": "Emirates", "iata": "EK", "icao": "UAE", "country": "UAE"},
        {"name": "Qatar Airways", "iata": "QR", "icao": "QTR", "country": "Qatar"},
        {"name": "Etihad Airways", "iata": "EY", "icao": "ETD", "country": "UAE"},
        {"name": "Turkish Airlines", "iata": "TK", "icao": "THY", "country": "Turkey"},
        {"name": "Saudi Arabian Airlines", "iata": "SV", "icao": "SVA", "country": "Saudi Arabia"},
        
        # Aerolíneas latinoamericanas
        {"name": "LATAM Airlines", "iata": "LA", "icao": "LAN", "country": "Chile"},
        {"name": "Avianca", "iata": "AV", "icao": "AVA", "country": "Colombia"},
        {"name": "Copa Airlines", "iata": "CM", "icao": "CMP", "country": "Panama"},
        {"name": "Aeroméxico", "iata": "AM", "icao": "AMX", "country": "Mexico"},
        {"name": "Volaris", "iata": "Y4", "icao": "VOI", "country": "Mexico"},
        {"name": "Interjet", "iata": "4O", "icao": "ABC", "country": "Mexico"},
        {"name": "JetSMART", "iata": "JA", "icao": "JAT", "country": "Chile"},
        {"name": "GOL Linhas Aéreas", "iata": "G3", "icao": "GLO", "country": "Brazil"},
        {"name": "Azul Brazilian Airlines", "iata": "AD", "icao": "AZU", "country": "Brazil"},
        {"name": "TAM Airlines", "iata": "JJ", "icao": "TAM", "country": "Brazil"},
        
        # Aerolíneas low-cost globales
        {"name": "Spirit Airlines", "iata": "NK", "icao": "NKS", "country": "USA"},
        {"name": "Frontier Airlines", "iata": "F9", "icao": "FFT", "country": "USA"},
        {"name": "Allegiant Air", "iata": "G4", "icao": "AAY", "country": "USA"},
        {"name": "Wizz Air", "iata": "W6", "icao": "WZZ", "country": "Hungary"},
        {"name": "Vueling", "iata": "VY", "icao": "VLG", "country": "Spain"},
        {"name": "Norwegian Air", "iata": "DY", "icao": "NAX", "country": "Norway"},
        {"name": "Pegasus Airlines", "iata": "PC", "icao": "PGT", "country": "Turkey"},
        
        # Aerolíneas de carga
        {"name": "FedEx Express", "iata": "FX", "icao": "FDX", "country": "USA"},
        {"name": "UPS Airlines", "iata": "5X", "icao": "UPS", "country": "USA"},
        {"name": "DHL Aviation", "iata": "D0", "icao": "DHX", "country": "Germany"},
        {"name": "Atlas Air", "iata": "5Y", "icao": "GTI", "country": "USA"},
        {"name": "Cargolux", "iata": "CV", "icao": "CLX", "country": "Luxembourg"},
    ]
    
    results = []
    query_upper = query.upper()
    
    # Buscar por nombre, código IATA o ICAO
    for airline in aerolineas_db:
        match_found = False
        match_score = 0
        
        # Búsqueda exacta en códigos tiene prioridad
        if query_upper == airline["iata"] or query_upper == airline["icao"]:
            match_found = True
            match_score = 100
        # Búsqueda parcial en nombre
        elif query_upper in airline["name"].upper():
            match_found = True
            match_score = 80
        # Búsqueda parcial en códigos
        elif query_upper in airline["iata"] or query_upper in airline["icao"]:
            match_found = True
            match_score = 60
        
        if match_found:
            display_name = f"{airline['name']} ({airline['iata']})"
            
            results.append({
                'id': airline["iata"].lower(),
                'name': display_name,
                'type': 'airline',
                'iata_code': airline["iata"],
                'icao_code': airline["icao"],
                'country': airline["country"],
                'match_score': match_score
            })
    
    # Ordenar por relevancia (match_score)
    results.sort(key=lambda x: x.get('match_score', 0), reverse=True)
    
    # Si no se encontró nada, crear sugerencias genéricas
    if not results and len(query) >= 2:
        results.append({
            'id': f'{query.lower()}_generic',
            'name': f'{query.upper()} Airlines',
            'type': 'airline',
            'generic': True
        })
    
    return results


@require_http_methods(["GET"])
def buscar_navios(request):
    """
    Buscar navíos con fallback mejorado
    """
    query = request.GET.get('q', '').strip()
    
    if len(query) < 3:
        return JsonResponse({'results': []})
    
    try:
        print(f"Buscando navíos: {query}")
        
        # Intentar con MyShipTracking primero
        results = buscar_navios_myshiptracking(query)
        
        if results:
            return JsonResponse({
                'results': results[:10],
                'source': 'MyShipTracking API',
                'total': len(results)
            })
        
        # Si no hay resultados de la API, usar fallback
        results = buscar_navios_fallback(query)
        
        return JsonResponse({
            'results': results[:10],
            'source': 'Fallback Database',
            'total': len(results)
        })
        
    except Exception as e:
        print(f"Error general en buscar_navios: {e}")
        # En caso de error, usar fallback
        results = buscar_navios_fallback(query)
        return JsonResponse({
            'results': results[:10],
            'source': 'Fallback (Error)',
            'error': str(e)
        })


def buscar_navios_myshiptracking(query):
    """
    MyShipTracking API con URL corregida y mejor manejo de errores
    """
    try:
        api_key = "xsbnnhmmZ8$lXDqEX6u7FQKXsJmtN8fqKA"
        secret_key = "DNfJ0Z7tF"
        
        # URLs posibles - prueba estas en orden:
        urls_posibles = [
            "https://api.myshiptracking.com/api/v1/vessels/search",
            "https://api.myshiptracking.com/v1/search",
            "https://api.myshiptracking.com/search",
            "https://myshiptracking.com/api/vessels/search",
        ]
        
        headers = {
            'X-API-Key': api_key,
            'X-Secret-Key': secret_key,
            'Content-Type': 'application/json',
            'User-Agent': 'Mozilla/5.0 (compatible; ShipTracker/1.0)'
        }
        
        params = {
            'name': query,
            'limit': 15
        }
        
        # Probar cada URL hasta encontrar una que funcione
        for url in urls_posibles:
            try:
                print(f"Probando URL: {url}")
                response = requests.get(url, headers=headers, params=params, timeout=8)
                
                if response.status_code == 200:
                    print(f"✅ URL funcional encontrada: {url}")
                    data = response.json()
                    results = []
                    
                    vessels = data.get('data', []) or data.get('vessels', []) or data.get('results', [])
                    
                    for vessel in vessels:
                        vessel_name = vessel.get('name') or vessel.get('ship_name') or vessel.get('vessel_name', '')
                        imo = vessel.get('imo') or vessel.get('imo_number', '')
                        mmsi = vessel.get('mmsi') or vessel.get('mmsi_number', '')
                        vessel_type = vessel.get('type') or vessel.get('ship_type', '')
                        
                        if vessel_name:
                            display_name = vessel_name
                            if imo:
                                display_name += f" (IMO: {imo})"
                            elif mmsi:
                                display_name += f" (MMSI: {mmsi})"
                            
                            results.append({
                                'id': imo or mmsi or vessel_name.replace(' ', '_').lower(),
                                'name': display_name,
                                'type': 'ship',
                                'imo': imo,
                                'mmsi': mmsi,
                                'vessel_type': vessel_type
                            })
                    
                    return results
                    
                elif response.status_code == 404:
                    continue
                else:
                    continue
                    
            except requests.exceptions.ConnectionError:
                continue
            except Exception as e:
                continue
        
        return []
            
    except Exception as e:
        return []


# Alternativa: usar una API diferente como VesselFinder o MarineTraffic
def buscar_navios_alternativa(query):
    """
    Alternativa usando una API pública diferente
    """
    try:
        # Ejemplo con VesselFinder API (necesitas registrarte para obtener API key)
        # url = "https://api.vesselfinder.com/vessels"
        
        # Por ahora, usar solo el fallback mejorado
        print("Usando fallback en lugar de API externa")
        return []
        
    except Exception as e:
        print(f"Error API alternativa: {e}")
        return []

def buscar_navios_fallback(query):
    """
    Base de datos expandida con nombres de barcos reales para autocompletar
    """
    navios_db = [
        # === PORTACONTENEDORES FAMOSOS ===
        {"name": "EVER GIVEN", "imo": "9811000", "type": "Container Ship", "company": "Evergreen Marine", "flag": "Panama"},
        {"name": "EVER GOLDEN", "imo": "9811012", "type": "Container Ship", "company": "Evergreen Marine", "flag": "Panama"},
        {"name": "EVER GLOBE", "imo": "9811024", "type": "Container Ship", "company": "Evergreen Marine", "flag": "Panama"},
        {"name": "MSC OSCAR", "imo": "9811046", "type": "Container Ship", "company": "MSC", "flag": "Panama"},
        {"name": "MSC ZARA", "imo": "9619881", "type": "Container Ship", "company": "MSC", "flag": "Liberia"},
        {"name": "MSC DIANA", "imo": "9621465", "type": "Container Ship", "company": "MSC", "flag": "Liberia"},
        {"name": "MSC MAYA", "imo": "9619893", "type": "Container Ship", "company": "MSC", "flag": "Liberia"},
        {"name": "MAERSK MADRID", "imo": "9778425", "type": "Container Ship", "company": "Maersk Line", "flag": "Denmark"},
        {"name": "MAERSK MILAN", "imo": "9778437", "type": "Container Ship", "company": "Maersk Line", "flag": "Denmark"},
        {"name": "MAERSK MUNICH", "imo": "9778449", "type": "Container Ship", "company": "Maersk Line", "flag": "Denmark"},
        {"name": "MAERSK NEBRASKA", "imo": "9778451", "type": "Container Ship", "company": "Maersk Line", "flag": "Denmark"},
        {"name": "CMA CGM MARCO POLO", "imo": "9454436", "type": "Container Ship", "company": "CMA CGM", "flag": "France"},
        {"name": "CMA CGM BENJAMIN FRANKLIN", "imo": "9745729", "type": "Container Ship", "company": "CMA CGM", "flag": "France"},
        {"name": "CMA CGM ANTOINE DE SAINT EXUPERY", "imo": "9745731", "type": "Container Ship", "company": "CMA CGM", "flag": "France"},
        {"name": "COSCO SHIPPING UNIVERSE", "imo": "9795592", "type": "Container Ship", "company": "COSCO", "flag": "China"},
        {"name": "COSCO SHIPPING GALAXY", "imo": "9795606", "type": "Container Ship", "company": "COSCO", "flag": "China"},
        {"name": "COSCO SHIPPING SOLAR", "imo": "9795618", "type": "Container Ship", "company": "COSCO", "flag": "China"},
        {"name": "OOCL HONG KONG", "imo": "9833910", "type": "Container Ship", "company": "OOCL", "flag": "Hong Kong"},
        {"name": "OOCL MADRID", "imo": "9833922", "type": "Container Ship", "company": "OOCL", "flag": "Hong Kong"},
        {"name": "OOCL SHENZHEN", "imo": "9833934", "type": "Container Ship", "company": "OOCL", "flag": "Hong Kong"},
        
        # === CRUCEROS FAMOSOS ===
        {"name": "SYMPHONY OF THE SEAS", "imo": "9744001", "type": "Cruise Ship", "company": "Royal Caribbean", "flag": "Bahamas"},
        {"name": "HARMONY OF THE SEAS", "imo": "9692596", "type": "Cruise Ship", "company": "Royal Caribbean", "flag": "Bahamas"},
        {"name": "ALLURE OF THE SEAS", "imo": "9398124", "type": "Cruise Ship", "company": "Royal Caribbean", "flag": "Bahamas"},
        {"name": "OASIS OF THE SEAS", "imo": "9398112", "type": "Cruise Ship", "company": "Royal Caribbean", "flag": "Bahamas"},
        {"name": "WONDER OF THE SEAS", "imo": "9863819", "type": "Cruise Ship", "company": "Royal Caribbean", "flag": "Bahamas"},
        {"name": "NORWEGIAN BLISS", "imo": "9751509", "type": "Cruise Ship", "company": "Norwegian Cruise Line", "flag": "Bahamas"},
        {"name": "NORWEGIAN EPIC", "imo": "9410569", "type": "Cruise Ship", "company": "Norwegian Cruise Line", "flag": "Bahamas"},
        {"name": "CARNIVAL VISTA", "imo": "9700424", "type": "Cruise Ship", "company": "Carnival Cruise Line", "flag": "Panama"},
        {"name": "CARNIVAL HORIZON", "imo": "9781024", "type": "Cruise Ship", "company": "Carnival Cruise Line", "flag": "Panama"},
        {"name": "MSC SEASIDE", "imo": "9706722", "type": "Cruise Ship", "company": "MSC Cruises", "flag": "Malta"},
        {"name": "MSC MERAVIGLIA", "imo": "9744101", "type": "Cruise Ship", "company": "MSC Cruises", "flag": "Malta"},
        {"name": "CELEBRITY EDGE", "imo": "9781999", "type": "Cruise Ship", "company": "Celebrity Cruises", "flag": "Malta"},
        
        # === TANQUEROS PETROLEROS ===
        {"name": "SEAWISE GIANT", "imo": "7381154", "type": "Oil Tanker", "company": "Retired", "flag": "Singapore"},
        {"name": "TI EUROPE", "imo": "9213891", "type": "Oil Tanker", "company": "Euronav", "flag": "Belgium"},
        {"name": "TI OCEANIA", "imo": "9213883", "type": "Oil Tanker", "company": "Euronav", "flag": "Belgium"},
        {"name": "TI ASIA", "imo": "9213907", "type": "Oil Tanker", "company": "Euronav", "flag": "Belgium"},
        {"name": "FRONT ALTAIR", "imo": "9253830", "type": "Oil Tanker", "company": "Frontline", "flag": "Marshall Islands"},
        {"name": "FRONT COMMANDER", "imo": "9156123", "type": "Oil Tanker", "company": "Frontline", "flag": "Marshall Islands"},
        {"name": "NORDIC SPACE", "imo": "9395220", "type": "Oil Tanker", "company": "Nordic Tankers", "flag": "Denmark"},
        {"name": "ATLANTIC EXPLORER", "imo": "9247465", "type": "Oil Tanker", "company": "Atlantic Tankers", "flag": "Liberia"},
        
        # === BULK CARRIERS (GRANELEROS) ===
        {"name": "BERGE STAHL", "imo": "8022915", "type": "Bulk Carrier", "company": "Berge Bulk", "flag": "Norway"},
        {"name": "VALE BRASIL", "imo": "9545647", "type": "Bulk Carrier", "company": "Vale", "flag": "Brazil"},
        {"name": "VALE RIO DE JANEIRO", "imo": "9545659", "type": "Bulk Carrier", "company": "Vale", "flag": "Brazil"},
        {"name": "PACIFIC CHALLENGER", "imo": "9384890", "type": "Bulk Carrier", "company": "Pacific Basin", "flag": "Hong Kong"},
        {"name": "CAPE FLORES", "imo": "9427834", "type": "Bulk Carrier", "company": "Capesize Shipping", "flag": "Marshall Islands"},
        {"name": "IRON CHIEFTAIN", "imo": "9398765", "type": "Bulk Carrier", "company": "Iron Ore Transport", "flag": "Liberia"},
        
        # === BARCOS HISTÓRICOS FAMOSOS ===
        {"name": "TITANIC", "imo": "Historical", "type": "Passenger Ship", "company": "White Star Line", "flag": "UK"},
        {"name": "QUEEN MARY 2", "imo": "9241061", "type": "Ocean Liner", "company": "Cunard Line", "flag": "UK"},
        {"name": "QUEEN ELIZABETH", "imo": "9477438", "type": "Cruise Ship", "company": "Cunard Line", "flag": "UK"},
        {"name": "QUEEN VICTORIA", "imo": "9320556", "type": "Cruise Ship", "company": "Cunard Line", "flag": "UK"},
        
        # === FERRY Y RORO ===
        {"name": "STENA HOLLANDICA", "imo": "9419567", "type": "RoRo Ferry", "company": "Stena Line", "flag": "Netherlands"},
        {"name": "PRIDE OF HULL", "imo": "9156147", "type": "RoRo Ferry", "company": "P&O Ferries", "flag": "UK"},
        {"name": "SILJA SERENADE", "imo": "8902012", "type": "RoRo Ferry", "company": "Tallink Silja", "flag": "Finland"},
        {"name": "COLOR FANTASY", "imo": "9306354", "type": "RoRo Ferry", "company": "Color Line", "flag": "Norway"},
        
        # === CARGUEROS GENERALES ===
        {"name": "ATLANTIC CARTIER", "imo": "9245678", "type": "General Cargo", "company": "Atlantic Shipping", "flag": "Canada"},
        {"name": "MEDITERRANEAN STAR", "imo": "9123456", "type": "General Cargo", "company": "Mediterranean Lines", "flag": "Italy"},
        {"name": "NORDIC TRADER", "imo": "9234567", "type": "General Cargo", "company": "Nordic Shipping", "flag": "Sweden"},
        {"name": "BALTIC PRINCESS", "imo": "9345678", "type": "General Cargo", "company": "Baltic Lines", "flag": "Estonia"},
        
        # === BARCOS ESPAÑOLES Y LATINOAMERICANOS ===
        {"name": "CIUDAD DE VALENCIA", "imo": "9456789", "type": "Container Ship", "company": "Naviera Española", "flag": "Spain"},
        {"name": "BARCELONA EXPRESS", "imo": "9567890", "type": "Container Ship", "company": "Mediterranean Shipping", "flag": "Spain"},
        {"name": "CRISTÓBAL COLÓN", "imo": "9678901", "type": "Cruise Ship", "company": "Pullmantur", "flag": "Spain"},
        {"name": "SANTÍSIMA TRINIDAD", "imo": "Historical", "type": "Naval Ship", "company": "Spanish Navy", "flag": "Spain"},
        {"name": "DON JUAN DE AUSTRIA", "imo": "9789012", "type": "Naval Ship", "company": "Spanish Navy", "flag": "Spain"},
        {"name": "PATAGONIA", "imo": "9890123", "type": "General Cargo", "company": "Naviera Argentina", "flag": "Argentina"},
        {"name": "MAGELLAN STAR", "imo": "9901234", "type": "Container Ship", "company": "Chilean Shipping", "flag": "Chile"},
        {"name": "AMAZONAS", "imo": "9012345", "type": "Bulk Carrier", "company": "Brazilian Shipping", "flag": "Brazil"},
        
        # === NOMBRES GENÉRICOS PARA COMPLETAR ===
        {"name": "ATLANTIC PRIDE", "type": "Container Ship", "company": "Atlantic Lines"},
        {"name": "PACIFIC STAR", "type": "Bulk Carrier", "company": "Pacific Shipping"},
        {"name": "MEDITERRANEAN QUEEN", "type": "Cruise Ship", "company": "Med Cruises"},
        {"name": "NORTHERN LIGHT", "type": "Oil Tanker", "company": "Northern Shipping"},
        {"name": "SOUTHERN CROSS", "type": "General Cargo", "company": "Southern Lines"},
        {"name": "EASTERN SPIRIT", "type": "Container Ship", "company": "Eastern Shipping"},
        {"name": "WESTERN WIND", "type": "Bulk Carrier", "company": "Western Lines"},
        {"name": "GLOBAL NAVIGATOR", "type": "Container Ship", "company": "Global Shipping"},
        {"name": "OCEAN VOYAGER", "type": "General Cargo", "company": "Ocean Lines"},
        {"name": "SEA CHAMPION", "type": "Oil Tanker", "company": "Sea Transport"},
    ]
    
    results = []
    query_upper = query.upper()
    
    # Buscar coincidencias exactas y parciales
    for navio in navios_db:
        nombre_match = query_upper in navio["name"].upper()
        company_match = navio.get("company") and query_upper in navio["company"].upper()
        type_match = query_upper in navio["type"].upper()
        
        if nombre_match or company_match or type_match:
            display_name = navio["name"]
            
            # Agregar información adicional
            info_parts = []
            if navio.get("imo") and navio["imo"] != "Historical":
                info_parts.append(f"IMO: {navio['imo']}")
            if navio.get("company"):
                info_parts.append(navio['company'])
            if navio.get("flag"):
                info_parts.append(f"Flag: {navio['flag']}")
            
            if info_parts:
                display_name += f" ({', '.join(info_parts[:2])})"  # Máximo 2 elementos
            
            results.append({
                'id': navio.get("imo", navio["name"].replace(' ', '_').lower()),
                'name': display_name,
                'type': 'ship',
                'imo': navio.get("imo", ""),
                'vessel_type': navio["type"],
                'company': navio.get("company", ""),
                'flag': navio.get("flag", "")
            })
    
    # Si no se encontró nada específico, generar sugerencias basadas en la query
    if not results and len(query) >= 2:
        tipos_barco = [
            ("Container Ship", "Portacontenedores"),
            ("Bulk Carrier", "Granelero"), 
            ("Oil Tanker", "Petrolero"),
            ("General Cargo", "Carga General"),
            ("Cruise Ship", "Crucero")
        ]
        
        for i, (tipo_en, tipo_es) in enumerate(tipos_barco[:3]):
            results.append({
                'id': f'{query.lower().replace(" ", "_")}_{i}',
                'name': f'{query.upper()} ({tipo_es})',
                'type': 'ship',
                'vessel_type': tipo_en,
                'generic': True,
                'suggestion': True
            })
    
    return results


# Simplificar la función principal
def buscar_navios_simplificado(request):
    """
    Versión simplificada que solo usa la base de datos local
    """
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({
            'results': [],
            'message': 'Consulta muy corta (mínimo 2 caracteres)'
        })
    
    try:
        print(f"🔍 Buscando navíos: '{query}'")
        
        # Usar solo el fallback (base de datos local)
        results = buscar_navios_fallback(query)
        
        print(f"📚 Encontrados {len(results)} resultados")
        
        return JsonResponse({
            'results': results[:15],  # Máximo 15 resultados
            'source': 'Local Ship Database',
            'total': len(results),
            'query': query
        })
        
    except Exception as e:
        print(f"❌ Error en búsqueda de navíos: {e}")
        return JsonResponse({
            'results': [],
            'error': str(e),
            'query': query
        })

@require_http_methods(["GET"])
def buscar_transporte(request):
    """
    Endpoint unificado que busca según el tipo de transporte con mejor debugging
    """
    tipo_transporte = request.GET.get('tipo', '').lower()
    query = request.GET.get('q', '').strip()
    
    print(f"� Búsqueda de transporte - Tipo: '{tipo_transporte}', Query: '{query}'")
    
    # Validar parámetros
    if not tipo_transporte:
        return JsonResponse({
            'results': [],
            'error': 'Parámetro "tipo" requerido',
            'valid_types': ['aereo', 'maritimo']
        })
    
    if not query:
        return JsonResponse({
            'results': [],
            'error': 'Parámetro "q" (query) requerido'
        })
    
    if tipo_transporte in ['aereo', 'aéreo']:
        print("✈️ Redirigiendo a búsqueda aérea...")
        return buscar_aeronaves(request)
    elif tipo_transporte in ['maritimo', 'marítimo']:
        print("🚢 Redirigiendo a búsqueda marítima...")
        return buscar_navios(request)
    else:
        print(f"❌ Tipo de transporte no válido: '{tipo_transporte}'")
        return JsonResponse({
            'results': [],
            'error': f'Tipo de transporte no válido: {tipo_transporte}',
            'valid_types': ['aereo', 'maritimo'],
            'received_type': tipo_transporte
        })
    
def prueba_envio_correo(request):
    send_mail(
        subject='Prueba de envío',
        message='Este es un correo de prueba desde Django.',
        from_email='hanswoods96@gmail.com',
        recipient_list=['hans.arancibia@live.com'],
        fail_silently=False,
    )
    return HttpResponse('Correo enviado correctamente.')


@login_required
@user_passes_test(lambda u: u.is_superuser)
def vista_folios_disponibles(request):
    from .models import Factura

    folio_min = 545
    folio_max = 10000
    usados = set(
        Factura.objects
        .exclude(folio_sii__isnull=True)
        .values_list('folio_sii', flat=True)
    )

    todos = list(range(folio_min, folio_max + 1))
    disponibles = sorted([f for f in todos if f not in usados])

    context = {
        'folio_min': folio_min,
        'folio_max': folio_max,
        'total_folios': folio_max - folio_min + 1,
        'folios_usados': sorted(usados),
        'folios_disponibles': disponibles,
        'cantidad_disponibles': len(disponibles),
        'cantidad_usados': len(usados),
    }

    return render(request, 'core/folios_disponibles.html', context)


@user_passes_test(lambda u: u.is_superuser)
def reenviar_factura_xml(request, factura_id):
    factura = get_object_or_404(Factura, id=factura_id)

    try:
        resultado = emitir_factura_exenta_cl_xml(factura)
        if resultado.get("success"):
            mensaje = "✅ XML reenviado correctamente"
        else:
            mensaje = f"❌ Error al reenviar: {resultado.get('error')}"

    except Exception as e:
        mensaje = f"🧨 Excepción: {str(e)}"

    return JsonResponse({"success": True, "mensaje": mensaje})


@user_passes_test(lambda u: u.is_superuser)
def descargar_factura_xml(request, factura_id):
    factura = get_object_or_404(Factura, id=factura_id)
    contenido = generar_xml_factura_exenta(factura)
    filename = f"factura_exenta_C{factura.certificado.id}.xml"
    response = HttpResponse(contenido, content_type='application/xml')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@login_required
def exportar_cobranzas_excel(request):
    from .models import Cobranza
    from decimal import Decimal

    qs = Cobranza.objects.select_related(
        'certificado',
        'certificado__cliente',
        'certificado__metodo_embarque',
        'certificado__factura',
        'certificado__notas'
    ).all()

    # Filtros (mismos de la vista)
    cliente = request.GET.get("cliente")
    rut = request.GET.get("rut")
    certificado = request.GET.get("certificado")
    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")

    if cliente:
        qs = qs.filter(certificado__cliente__nombre__icontains=cliente)
    if rut:
        qs = qs.filter(certificado__cliente__rut__icontains=rut)
    if certificado:
        qs = qs.filter(certificado__id__icontains=certificado)
    if inicio:
        qs = qs.filter(certificado__fecha_creacion__gte=inicio)
    if fin:
        qs = qs.filter(certificado__fecha_creacion__lte=fin)

    # ---------- Workbook ----------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Cobranzas"

    # Controles de reporte (editables por el admin)
    ws["A1"] = "Controles de reporte"; ws["A1"].font = Font(bold=True)
    ws["A2"] = "TC reporte (CLP/USD)"
    ws["A3"] = "Tasa override (%)"
    ws["A4"] = "Mínimo override (USD)"

    # TC por defecto: intenta usar el último tipo_cambio existente; si no, 950
    default_tc = 950.0
    for c in qs.order_by("-certificado__fecha_creacion"):
        fac = getattr(c.certificado, "factura", None)
        if fac and getattr(fac, "tipo_cambio", None):
            try:
                default_tc = float(fac.tipo_cambio)
                break
            except Exception:
                pass
    ws["B2"] = default_tc
    ws["B3"] = ""   # vacío = usa tasa del cliente
    ws["B4"] = ""   # vacío = usa mínimo del cliente

    start = 6  # fila de encabezados
    headers = [
        "N° Certificado","N° Factura","Cliente","RUT",
        "Valor Seguro (USD)","Valor Prima (USD)","Valor Factura (CLP)","Referencia",
        "Tasa %","Mínimo USD","Prima USD (calc)","A pagar a SyC (CLP)"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    r = start + 1
    for cobro in qs:
        cert = cobro.certificado
        fac  = cert.factura
        cli  = cert.cliente

        # columnas base
        ws.cell(row=r, column=1).value = f"C-{cert.id}"
        ws.cell(row=r, column=2).value = fac.numero if fac else ""
        ws.cell(row=r, column=3).value = cli.nombre
        ws.cell(row=r, column=4).value = cli.rut
        ws.cell(row=r, column=5).value = float(cobro.monto_asegurado or 0)          # Valor Seguro USD
        ws.cell(row=r, column=6).value = float(cobro.valor_prima_cobro or 0)        # Prima USD guardada
        ws.cell(row=r, column=7).value = float(fac.valor_clp) if fac and fac.valor_clp else 0  # CLP contable
        ws.cell(row=r, column=8).value = cert.notas.referencia if getattr(cert, "notas", None) and cert.notas.referencia else ""

        # tasa/minimo del cliente (ajusta nombres si en tu modelo son otros)
        tasa_cli   = float(getattr(cli, "tasa", 0) or 0)
        minimo_cli = float(getattr(cli, "minimo", 0) or 0)
        ws.cell(row=r, column=9).value  = tasa_cli
        ws.cell(row=r, column=10).value = minimo_cli

        # K: Prima USD (calc) = MAX(E*r/100, min) usando overrides si los pones
        ws.cell(row=r, column=11).value = (
            f'=ROUND(MAX(E{r}*IF($B$3<>"",$B$3,I{r})/100, IF($B$4<>"",$B$4,J{r})), 2)'
        )
        # L: A pagar a SyC (CLP) = Prima USD (calc) * TC reporte
        ws.cell(row=r, column=12).value = f'=ROUND(K{r}*$B$2, 0)'

        r += 1

    # formatos
    for row in ws.iter_rows(min_row=start+1, max_row=r-1, min_col=5, max_col=12):
        for idx, cell in enumerate(row, start=5):
            if idx in (5, 6, 11):  # USD
                cell.number_format = '#,##0.00'
            if idx in (7, 12):     # CLP
                cell.number_format = '#,##0'

    # anchos
    widths = [16,16,28,14,18,18,20,18,10,12,16,18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # congelar encabezados / auto-filtro
    ws.freeze_panes = ws["A7"]
    ws.auto_filter.ref = f"A{start}:L{r-1}"

    # respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_cobranzas.xlsx"'
    wb.save(response)
    return response



@login_required
@user_passes_test(lambda u: u.is_superuser)
def test_envio_certificado_real(request):
    try:
        certificado = CertificadoTransporte.objects.latest('id')
        pdf_cert = generar_pdf_certificado(certificado, request)
        pdf_fact = generar_pdf_factura(certificado, request)

        # Destinatarios de prueba
        destinatarios = ['hans.arancibia@live.com']  # agrega otros si quieres
        if certificado.creado_por and certificado.creado_por.correo:
            destinatarios.append(certificado.creado_por.correo)

        enviar_certificado_y_factura(certificado, pdf_cert, pdf_fact, destinatarios)

        return HttpResponse("✅ Correo de prueba con certificado y factura enviado correctamente.")
    except Exception as e:
        return HttpResponse(f"❌ Error al enviar correo de prueba: {str(e)}")
@csrf_exempt
def buscar_puertos_api_externa(request):
    term = (request.GET.get('term') or '').strip()
    if not term:
        return JsonResponse({'error': 'Término de búsqueda no proporcionado'}, status=400)

    # Opcionales para filtrar por país
    pais = (request.GET.get('pais') or '').strip()
    pais_code = (request.GET.get('pais_code') or '').strip()
    try:
        limit = int(request.GET.get('limit', 25))
    except ValueError:
        limit = 25

    # Normalización a ISO-2 (alias típicos para EE.UU.)
    aliases = {
        'united states': 'US',
        'united states of america': 'US',
        'usa': 'US',
        'u.s.': 'US',
        'u.s.a.': 'US',
    }
    iso2 = (pais_code.upper() if len(pais_code) == 2 else '') or aliases.get(pais.lower(), '')
    if not iso2 and len(pais) == 2:
        iso2 = pais.upper()

    # Conjunto de nombres aceptados para el filtro por nombre de país
    accepted_country_names = set()
    if pais:
        accepted_country_names.add(pais)
    if iso2 == 'US':
        accepted_country_names.update([
            'United States', 'United States of America', 'USA', 'U.S.', 'U.S.A.'
        ])

    api_url = f'https://igi.nsure.net/api/seaports/search/term/{quote(term)}'
    headers = {
        'Content-Type': 'application/json',
        'Accept': 'application/json',
        'X-NS-API-Key': 'PqrED4vo2UI8I8TlPTgOKmCo0C1OuSUzbgbpIBHQwgnA2343xbtwqjhmGZ2bckvp',
    }

    try:
        resp = requests.get(api_url, headers=headers, timeout=12)
        resp.raise_for_status()
        data = resp.json()

        # La API suele devolver una lista; si viniera envuelta, intenta sacar "results"
        items = data if isinstance(data, list) else data.get('results', [])

        # Filtro por país (si recibimos pais/pais_code)
        def matches_country(item):
            if not iso2 and not accepted_country_names:
                return True  # no hay filtro de país
            code = (item.get('countryCode') or item.get('country_code') or item.get('country') or '').strip()
            name = (item.get('countryName') or item.get('country_name') or item.get('country') or '').strip()
            if iso2 and code.upper() == iso2:
                return True
            if name and any(name.lower() == n.lower() for n in accepted_country_names):
                return True
            return False

        filtered = [it for it in items if matches_country(it)]

        # Fallback: si el filtro dejó cero y sí había resultados, entrega lo original (mejor UX)
        out = filtered if (filtered or not items) else items
        if limit and isinstance(out, list):
            out = out[:limit]

        return JsonResponse(out, safe=False)
    except requests.exceptions.RequestException as e:
        return JsonResponse({'error': 'No se pudo obtener la información', 'detalle': str(e)}, status=500)

def test_email_view(request):
    try:
        send_mail(
            'Asunto de Prueba desde Django y Mailgun API', # Asunto actualizado para Mailgun
            'Este es el cuerpo del mensaje de prueba enviado a través de Mailgun API.', # Cuerpo actualizado para Mailgun
            settings.DEFAULT_FROM_EMAIL, # Remitente: usará el correo configurado en DEFAULT_FROM_EMAIL (debe ser de tu dominio de Mailgun)
            ['hans.ti@safeyourcargo.com', 'finanzas@safeyourcargo.com'], # Destinatarios
            fail_silently=False,
        )
        # Mensaje de éxito actualizado para reflejar Mailgun
        return HttpResponse("Correo de prueba enviado con éxito con Mailgun a hans.ti@safeyourcargo.com y finanzas@safeyourcargo.com.")
    except Exception as e:
        # Mensaje de error actualizado para reflejar Mailgun
        return HttpResponse(f"Error al enviar el correo con Mailgun: {e}", status=500)
    

class NsureTestView(View):
    template_name = 'core/nsure_test.html'

    def get(self, request, *args, **kwargs):
        form = NsureTestForm()
        context = {
            'form': form,
            'api_response': None,
            'error_message': None,
            'created_declaration_id': None, # Para mostrar el ID de la declaración creada
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        form = NsureTestForm(request.POST)
        api_response_data = None
        error_message = None
        created_declaration_id = None

        if form.is_valid():
            endpoint_choice = form.cleaned_data['endpoint']
            search_term = form.cleaned_data['search_term']
            declaration_id = form.cleaned_data['declaration_id']

            try:
                if endpoint_choice == 'create_declaration':
                    # Genera un externalId único para cada declaración de prueba
                    current_time_str = datetime.now().strftime('%Y%m%d%H%M%S%f')
                    test_external_id = f"test_decl_{current_time_str}"
                    
                    # Llamar a la función para crear la declaración
                    response = nsure_api.create_declaration(
                        external_id=test_external_id, 
                        departure_date=datetime.now() # O una fecha específica si lo requiere
                    )
                    api_response_data = response
                    
                    # Extraer el ID o externalId de la respuesta
                    if response and 'externalId' in response:
                        created_declaration_id = f"external-id={response['externalId']}"
                    elif response and 'id' in response:
                        created_declaration_id = str(response['id'])
                    
                    if not created_declaration_id:
                        error_message = "Declaración creada, pero no se pudo extraer el ID de la respuesta."

                elif endpoint_choice == 'vessels':
                    if not declaration_id:
                        error_message = "Debes proporcionar un ID de Declaración para buscar navíos."
                    elif not search_term:
                        error_message = "Debes proporcionar un término de búsqueda para buscar navíos."
                    else:
                        api_response_data = nsure_api.search_vessels(declaration_id, search_term)
                        
                elif endpoint_choice == 'countries':
                    if not declaration_id:
                        error_message = "Debes proporcionar un ID de Declaración para listar países."
                    else:
                        api_response_data = nsure_api.list_countries(declaration_id)
                
            except Exception as e:
                # Captura el mensaje de error de la excepción
                error_message = f"Error al comunicarse con la API de NSure Cargo: {e}"
                # Si es un error HTTP, podrías intentar mostrar el cuerpo de la respuesta de error
                if hasattr(e, 'response') and e.response is not None:
                     error_message += f"\nDetalle: {e.response.text}"
        else:
            error_message = "Formulario inválido. Revisa los campos."
            # Mostrar errores específicos del formulario
            for field, errors in form.errors.items():
                error_message += f"\nCampo '{field}': {', '.join(errors)}"


        context = {
            'form': form,
            'api_response': json.dumps(api_response_data, indent=2) if api_response_data else None,
            'error_message': error_message,
            'created_declaration_id': created_declaration_id, # Pasa el ID al contexto
        }
        return render(request, self.template_name, context)



@require_http_methods(["GET"])
def buscar_aerolineas(request):
    """
    Vista para buscar aerolíneas en nuestra base de datos local.
    """
    query = request.GET.get('query', '')
    if len(query) < 2:
        return JsonResponse({'results': []})

    # Busca en la base de datos local coincidencias en el nombre O en el código IATA
    # 'icontains' es para búsqueda case-insensitive (no distingue mayúsculas/minúsculas)
    resultados = Aerolinea.objects.filter(
        Q(nombre__icontains=query) | 
        Q(codigo_iata__icontains=query)
    ).values('nombre', 'codigo_iata')[:10] # Limita a 10 resultados para ser rápido

    # Formatea la respuesta para que el JavaScript la entienda
    resultados_formateados = [
        {
            'name': f"{item['nombre']}", # Puedes ajustar cómo se ve el nombre
            'iata': item['codigo_iata']
        }
        for item in resultados
    ]
    
    return JsonResponse({'results': resultados_formateados})
def pagina_prueba_aeronaves(request):
    """
    Vista temporal para mostrar la página de prueba del autocompletado de aeronaves.
    """
    return render(request, 'core/prueba_aeronaves.html')


@require_http_methods(["GET"])
def buscar_navios(request):
    """
    Vista para buscar navíos en la base de datos local.
    """
    query = request.GET.get('query', '')
    if len(query) < 2:
        return JsonResponse({'results': []})

    # Busca en el modelo Navio por nombre o IMO
    resultados = Navio.objects.filter(
        Q(nombre__icontains=query) | 
        Q(imo__icontains=query)
    ).values('nombre', 'imo', 'naviera')[:15]

    # Formatea la respuesta para el JavaScript
    resultados_formateados = [
        {
            'name': f"{item['nombre']} ({item['naviera'] or 'N/A'})",
            'imo': item['imo']
        }
        for item in resultados
    ]
    
    return JsonResponse({'results': resultados_formateados})